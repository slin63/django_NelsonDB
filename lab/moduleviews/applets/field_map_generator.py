# Generates a field map when passed an appropriately formatted CSV with pre-filled row and range information
# www.github.com/slin63
# slin63@illinois.edu
from re import split
from openpyxl import Workbook
from openpyxl.utils import _get_column_letter
from openpyxl.styles import PatternFill, Border, Side
from datetime import datetime



class PlotCell(object):
    def __init__(self, range_num, row_num, experiment, plot_id, field):
        self.range = range_num
        self.row = row_num
        self.experiment = experiment
        self.plot_id = plot_id
        self.field = field

    def __repr__(self):
        return "Range {}, Row {}, Exp {} Pl_ID {} Field {}".format(
            self.range, self.row, self.experiment, self.plot_id, self.field
        )

    def get_row_list(self):
        return [int(e) for e in self.row.split('-')]

    def get_coordinates(self):
        if len(self.row) == 1:
            coordinates = ('{}{}'.format(self.range, self.row),)
        else:
            coordinates = ()
            row_list = self.get_row_list()
            for block in xrange(int(min(row_list)), int(max(row_list)) + 1):
                coordinates += ('{}{}'.format(self.range, block),)

        return coordinates


def compile_info(plot_objects, field=None):
    """
    :plot_objects: List containing selected ObsPlot objects
    :return: .xlsx containing a colored field map of the passed plots.
    """
    wb = Workbook()
    if field:
        field_set = [field]
    else:
        field_set = get_field_object_set(plot_objects)
    print 'field_set =', field_set
    for field in field_set:
        field_plots = get_plots_in_field(plot_objects, field)
        print 'field_plots = ', field_plots
        worksheet = wb.create_sheet()
        worksheet.title = field.field_name
        experiment_current = field_plots[0].experiment
        experiment_colors = iter(['00ccff', '00ffcc', 'ffcccc'])
        cell_color = experiment_colors.next()

        for plot in field_plots:
            if plot.experiment != experiment_current:
                experiment_current = plot.experiment
                # Cycles through three experiment colors so the spreadsheet doesn't look incredibly dull
                try:
                    cell_color = experiment_colors.next()
                except StopIteration:
                    experiment_colors = iter(['00ccff', '00ffcc', 'ffcccc'])
                    cell_color = experiment_colors.next()

            cell_fill = PatternFill(start_color=cell_color,
                       end_color=cell_color,
                       fill_type='solid')
            print plot
            print plot.get_coordinates()
            side = Side(border_style='thin', color="003300")
            for coordinate in plot.get_coordinates():
                worksheet[coordinate] = plot.plot_id
                worksheet[coordinate].fill = cell_fill
                border = Border(bottom=worksheet[coordinate].border.bottom)
                if check_row_even(coordinate): # Slips in a border every 2 rows
                    border.bottom = side
                    worksheet[coordinate].border = border


        domain = get_plot_domains(field_plots)

        add_axes(worksheet, domain, field_plots)

    wb.remove_sheet(wb.get_sheet_by_name('Sheet'))

    return wb


def add_axes(worksheet, domain, plot_objects):
    """
    :param worksheet: Worksheet we will be appending with information.
    :param domain: Rows and ranges.
    :return: Excel file with row and range axes.
    """
    axes = generate_axes(domain, plot_objects)
    for axis in axes:
        for coordinate in axis.keys():
            worksheet[coordinate] = axis[coordinate]

    return 0


def generate_axes(domain, plot_objects):
    """
    :param domain: Rows and ranges.
    :return: Dictionaries formatted {ExcelIndex (e.g. H23): Row or range value} to use as axes.
    """
    left_row = True
    top_range = True

    row_max = max(domain[0])
    row_min = min(domain[0])

    ranges = [letter_to_number(e) for e in domain[1]]
    range_max = _get_column_letter(max(ranges))
    range_min = _get_column_letter(min(ranges))

    range_max_plus_one = _get_column_letter(letter_to_number(range_max) + 1)
    try:
        range_min_sub_one = _get_column_letter(letter_to_number(range_min) - 1)
    except ValueError:
        range_min_sub_one = range_max_plus_one
        top_range = False

    row_max_plus_one = row_max + 1
    row_min_sub_one = row_min - 1
    if row_min_sub_one == 0:
        row_min_sub_one += 1
        left_row = False

    # Generate basic labels
    labels = {}
    if top_range:
        labels = {range_min_sub_one + str(row_min_sub_one): 'Rows/Ranges'}

    # Generate row axes labels
    row_axes = {}
    for e in xrange(row_min, row_max + 1):
        if left_row:
            row_axes[range_min_sub_one + str(e)] = e
        row_axes[range_max_plus_one + str(e)] = e

    # Generate range axes labels
    range_axes = {}
    for e in xrange(letter_to_number(range_min), letter_to_number(range_max) + 1):
        if top_range:
            range_axes[_get_column_letter(e) + str(row_min_sub_one)] = e
        range_axes[_get_column_letter(e) + str(row_max_plus_one)] = e

    # Generate experiment/script info labels
    experiments = get_plot_experiments(plot_objects)
    experiment_axes = {
        range_min + str(row_max + 2): 'FieldMapper / slin63@illinois.edu / FCPathology / Rendered: {}. Experiments described at bottom of sheet.'.format(datetime.now())
    }
    current_row = row_max + 3
    for exp in experiments:
        exp_string = 'EXP: {} - {}. Owner: {}. Field: {}. Purpose: {}. Comments: {}.'.format(exp.name, exp.start_date, exp.user, exp.field, exp.purpose, exp.comments)
        experiment_axes[range_min + str(current_row)] = exp_string
        current_row += 1



    return row_axes, range_axes, labels, experiment_axes


def empty_field():
    wb = Workbook()
    ws = wb.active
    ws['C3'] = 'No data for this field.'

    return wb

def letter_to_number(letter):
    letter = letter.lower()
    l_to_n = {
        'a': 1, 'b': 2, 'c': 3, 'd': 4, 'e': 5, 'f': 6, 'g': 7, 'h': 8, 'i': 9, 'j': 10, 'k': 11, 'l': 12, 'm': 13,
        'n': 14, 'o': 15, 'p': 16, 'q': 17, 'r': 18, 's': 19, 't': 20, 'u': 21, 'v': 22, 'w': 23, 'x': 24, 'y': 25,
        'z': 26, 'aa': 27, 'ab': 28, 'ac': 29, 'ad': 30, 'ae': 31, 'af': 32, 'ag': 33, 'ah': 34, 'ai': 35, 'aj': 36,
        'ak': 37, 'al': 38, 'am': 39, 'an': 40
    }
    return l_to_n[letter]


def get_field_object_set(object_list):
    field_set = set()
    for obj in object_list:
        field_set.add(obj.field)

    return field_set


def get_plots_in_field(object_list, field):
    plots_in_field = []
    for obj in object_list:
        if obj.field == field:
            plots_in_field.append(obj)

    return plots_in_field


def get_plot_domains(object_list):
    rows = []
    ranges = []
    for obj in object_list:
        rows += obj.get_row_list()
        ranges.append(obj.range)

    return [rows, ranges]


def get_plot_experiments(object_list):
    experiment_set = set()
    for obj in object_list:
        experiment_set.add(obj.experiment)

    return experiment_set


def check_row_even(coordinate):
    """
    Given a string of the form 'COLUMN-ROW' (e.g. 'AD45'), returns whether or not the row is even. (e.g. False from 'AD45')
    """
    return int(split('(\d+)', coordinate)[1]) % 2 == 0
