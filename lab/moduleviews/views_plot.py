import csv
import datetime

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import render_to_response
from django.template import RequestContext
from django.db import transaction
from itertools import chain
from openpyxl.writer.excel import save_virtual_workbook

from applets import field_map_generator
from lab.forms import DownloadFieldForm, HarvestDateForm, UpdatePlotsOnlineForm, TreatmentForm
from lab.models import Experiment, Stock, ObsPlot, ObsPlant, ObsSample, ObsWell, ObsCulture, ObsTissue, ObsDNA, \
    ObsPlate, ObsMicrobe, ObsExtract, ObsTracker, ObsTrackerSource, IsolateStock, Field, \
    Measurement, MaizeSample, Isolate, Treatment, MeasurementParameter


YEAR_INIT = 2015


@login_required
def add_harvest_date(request):
    context = RequestContext(request)
    context_dict = {}
    context_dict['logged_in_user'] = request.user.username
    view = None

    if request.method == 'POST':
        form = HarvestDateForm(request.POST, request.FILES)
        if form.is_valid():
            upload_file = request.FILES['file_name']
            if process_harvest_dates(upload_file):
                context_dict['success'] = "Harvest dates added!"
                context_dict['upload_form'] = HarvestDateForm()
                view = render_to_response("lab/plot/harvest_date.html", context_dict, context)
            else:
                context_dict['errors'] = "Invalid Form"
                context_dict['upload_form'] = HarvestDateForm()
                view = render_to_response("lab/plot/harvest_date.html", context_dict, context)

        else:
            context_dict['errors'] = "Incomplete Form"
            context_dict['upload_form'] = HarvestDateForm()
            view = render_to_response("lab/plot/harvest_date.html", context_dict, context)

    elif request.method == 'GET':
        form = HarvestDateForm()
        context_dict['upload_form'] = form
        view = render_to_response("lab/plot/harvest_date.html", context_dict, context)

    return view


@login_required
def add_treatment(request):
    context = RequestContext(request)
    context_dict = {}
    context_dict['logged_in_user'] = request.user.username
    view = None

    if request.method == 'POST':
        form = TreatmentForm(request.POST, request.FILES)
        if form.is_valid():
            upload_file = request.FILES['file_name']
            if process_treatments(upload_file):
                context_dict['success'] = "Treatments added!"
                context_dict['upload_form'] = TreatmentForm()
                view = render_to_response("lab/plot/add_treatment.html", context_dict, context)
            else:
                context_dict['errors'] = "Invalid Form"
                context_dict['upload_form'] = TreatmentForm()
                view = render_to_response("lab/plot/add_treatment.html", context_dict, context)

        else:
            context_dict['errors'] = "Incomplete Form"
            context_dict['upload_form'] = TreatmentForm()
            view = render_to_response("lab/plot/add_treatment.html", context_dict, context)

    elif request.method == 'GET':
        form = TreatmentForm()
        context_dict['upload_form'] = form
        view = render_to_response("lab/plot/add_treatment.html", context_dict, context)

    return view


def process_treatments(upload_file):
    success = True

    rdr = csv.DictReader(upload_file)
    try:
        for row in rdr:
            plot_id = row['Plot ID']
            treatment_id = row['Treatment ID']

            plot_obs = ObsTracker.objects.get(obs_entity_type='plot', obs_plot__plot_id=plot_id)
            treatment = Treatment.objects.get(treatment_id=treatment_id)

            plot_obs.obs_treatment = treatment
            plot_obs.save()

    except ObsTracker.DoesNotExist or Treatment.DoesNotExist:
        success = False
        pass

    return success


@login_required
def update_plot_info(request, obs_plot_id):
    context = RequestContext(request)
    context_dict = {}
    if request.method == 'POST':
        plot_form = UpdatePlotsOnlineForm(data=request.POST)
        if plot_form.is_valid():
            with transaction.atomic():
                try:
                    plot = ObsPlot.objects.get(id=obs_plot_id)
                    plot.plot_id = plot_form.cleaned_data['plot_id']
                    plot.plot_name = plot_form.cleaned_data['plot_name']
                    plot.polli_type = plot_form.cleaned_data['polli_type']
                    plot.gen = plot_form.cleaned_data['gen']
                    plot.shell_bulk = plot_form.cleaned_data['shell_bulk']
                    plot.shell_single = plot_form.cleaned_data['shell_single']
                    plot.is_male = plot_form.cleaned_data['is_male']
                    plot.cross_target = plot_form.cleaned_data['cross_target']
                    plot.save()
                    context_dict['updated'] = True
                except Exception:
                    context_dict['failed'] = True
        else:
            print(plot_form.errors)
    else:
        plot_data = ObsPlot.objects.filter(id=obs_plot_id).values('plot_id', 'plot_name', 'polli_type', 'gen', 'shell_bulk', 'shell_single', 'cross_target')
        try:
            plot_form = UpdatePlotsOnlineForm(initial=plot_data[0])
        except IndexError:
            plot_form = None

    try:
        context_dict['plot'] = ObsPlot.objects.get(id=obs_plot_id)
    except ObsPlot.DoesNotExist:
        context_dict['plot'] = None
    context_dict['obs_plot_id'] = obs_plot_id
    context_dict['plot_form'] = plot_form
    context_dict['logged_in_user'] = request.user.username
    return render_to_response('lab/plot/plot_info_update.html', context_dict, context)



def process_harvest_dates(upload_file):
    success = True
    # try:
    rdr = csv.DictReader(upload_file)
    try:
        for row in rdr:
            print row
            p_id = row["Plot ID"]
            date = row["Harvest Date"]
            plot = ObsPlot.objects.get(plot_id=p_id)
            plot.harvest_date = date
            plot.save()
    except ObsPlot.DoesNotExist:
        success = False
        pass

    return success


@login_required
def single_plot_info(request, obs_plot_id):
    context = RequestContext(request)
    context_dict = {}
    obs_tracker_plot = []
    try:
        plot_info = ObsPlot.objects.get(id=obs_plot_id)
    except ObsPlot.DoesNotExist:
        plot_info = None
        obs_measurements = None
    if plot_info is not None:
        obs_tracker_only_plot = ObsTracker.objects.get(obs_entity_type='plot', obs_plot_id=obs_plot_id)
        obs_tracker = get_obs_tracker('obs_plot_id', obs_plot_id)
        for t in obs_tracker:
            if t.obs_id != plot_info.plot_id:
                obs_tracker_plot.append(t)
        obs_source = get_seed_collected_from_plot('obs_plot_id', obs_plot_id)
        obs_measurements = get_obs_measurements('obs_plot_id', obs_plot_id)
    else:
        obs_tracker_plot = None
        obs_source = None
    context_dict['plot_obs_info'] = obs_tracker_only_plot
    context_dict['plot_info'] = plot_info
    context_dict['obs_tracker'] = obs_tracker_plot
    context_dict['obs_source'] = obs_source
    context_dict['obs_measurements'] = obs_measurements
    context_dict['logged_in_user'] = request.user.username

    return render_to_response('lab/plot/plot_info.html', context_dict, context)


@login_required
def plot_loader_from_experiment(request, experiment_name):
    context = RequestContext(request)
    context_dict = {}
    context_dict = checkbox_session_variable_check(request)
    plot_loader = find_plot_from_experiment(experiment_name)
    context_dict['plot_loader'] = plot_loader
    context_dict['experiment_name'] = experiment_name
    context_dict['logged_in_user'] = request.user.username
    return render_to_response('lab/plot/plot_experiment_data.html', context_dict, context)


@login_required
def download_plot_experiment(request, experiment_name):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="%s_plots.csv"' % (experiment_name)
    plot_loader = find_plot_from_experiment(experiment_name)
    writer = csv.writer(response)
    writer.writerow(
        ['Experiment Name', 'Plot ID', 'Plot Name', 'Polli_type', 'Shelling', 'Generation', 'Field', 'Source Stock', 'Row', 'Range', 'Plot', 'Block', 'Rep', 'Kernel Num',
         'Planting Date', 'Harvest Date', 'Comments'])
    for row in plot_loader:
        writer.writerow([row.experiment, row.obs_plot.plot_id, row.obs_plot.plot_name, row.obs_plot.polli_type, row.obs_plot.get_shell_type(), row.obs_plot.gen, row.field.field_name, row.stock.seed_id,
                         row.obs_plot.row_num, row.obs_plot.range_num, row.obs_plot.plot, row.obs_plot.block,
                         row.obs_plot.rep, row.obs_plot.kernel_num, row.obs_plot.planting_date,
                         row.obs_plot.harvest_date, row.obs_plot.comments])
    return response


@login_required
def download_plot_measurements(request, experiment_name):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="%s_plot_measurements.csv"' % (experiment_name)
    plot_loader = find_plot_from_experiment(experiment_name)
    measurements = Measurement.objects.filter(obs_tracker__obs_plot__plot_id__startswith=experiment_name)
    header = plot_loader[0].__dict__.keys() + measurements[0].__dict__.keys()
    header.remove('_state')

    fieldnames = [
        'plot_id',

        # Measurement fields
        'value', 'time_of_measurement', 'time_of_measurement_JULIAN', 'measurement_parameter_id', 'comments', 'experiment_id',

        # Plot fields
        'row_num', 'shell_multi', 'cross_target', 'polli_type', 'plot', 'rep', 'shell_bulk',
        'comments', 'gen', 'is_male', 'kernel_num', 'planting_date', 'planting_date_JULIAN', 'harvest_date',
        'plot_name', 'shell_single', 'range_num', 'block',
    ]

    writer = csv.DictWriter(response, fieldnames)
    writer.writeheader()

    for model in measurements:
        info = merge_dicts(model.__dict__, model.obs_tracker.obs_plot.__dict__)
        parameter_name = MeasurementParameter.objects.get(id=info['measurement_parameter_id']).parameter
        info['measurement_parameter_id'] = parameter_name
        info.pop('_state', None)
        info.pop('obs_tracker_id', None)
        info.pop('_obs_tracker_cache', None)
        info.pop('id', None)
        info.pop('user_id', None)

        info['time_of_measurement_JULIAN'] = model.get_julian(format="%m/%d/%Y %H:%M")
        info['planting_date_JULIAN'] = model.obs_tracker.obs_plot.get_julian(format="%m/%d/%Y")


        writer.writerow(info)

    return response


def merge_dicts(dict_a, dict_b):
    ret_dict = dict_a.copy()
    ret_dict.update(dict_b)
    return ret_dict


def find_plot_from_experiment(experiment_name):
    try:
        plot_loader = ObsTracker.objects.filter(experiment__name=experiment_name, obs_entity_type='plot')
    except ObsTracker.DoesNotExist:
        plot_loader = None
    return plot_loader


def get_fields_from_plots(object_list):
    field_set = set()
    for obj in object_list:
        field_set.add(obj.field)

    return field_set


@login_required
def plot_loader_browse(request):
    context = RequestContext(request)
    context_dict = {}

    # Handles the little select field-map form at the bottom
    if request.method == 'POST':
        # print request.POST
        form = DownloadFieldForm(request.POST)
        if form.is_valid():
            field_id = form.cleaned_data['field'].id
            if form.cleaned_data['get_csv_instead']:
                return download_plot_loader(request, field_id)
            else:
                return download_field_map_by_field(request, field_id)
                # return HttpResponseRedirect('/lab/download/plot_field/{}'.format(field_id))
    else:
        form = DownloadFieldForm()
        plot_loader = sort_plot_loader(request)
        field_list = Field.objects.exclude(id=1)
        context_dict = checkbox_session_variable_check(request)
        context_dict['form'] = form
        context_dict['plot_loader'] = plot_loader
        context_dict['field_list'] = field_list
        context_dict['logged_in_user'] = request.user.username
        return render_to_response('lab/plot/plot_data.html', context_dict, context)


def sort_plot_loader(request, field=None):
    plot_loader = {}
    if field:
        plot_loader = ObsTracker.objects.filter(field=field, obs_entity_type='plot')
    else:
        if request.session.get('checkbox_plot_experiment_id_list', None):
            checkbox_plot_experiment_id_list = request.session.get('checkbox_plot_experiment_id_list')
            for plot_experiment in checkbox_plot_experiment_id_list:
                rows = ObsTracker.objects.filter(obs_entity_type='plot', experiment__id=plot_experiment)
                plot_loader = list(chain(rows, plot_loader))
        else:
            plot_loader = ObsTracker.objects.filter(obs_entity_type='plot')[:100]
    return plot_loader


@login_required
def download_plot_loader(request, field_id=None):
    if field_id:
        field = Field.objects.get(id=field_id)
        filename = field.field_name.replace(' ', '') + '-Plots.csv'
        plot_loader = sort_plot_loader(request=request, field=field)
    else:
        filename = 'selected-experiments.csv'
        plot_loader = sort_plot_loader(request=request, field=None)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename={}'.format(filename)
    writer = csv.writer(response)
    writer.writerow(
        ['Experiment Name', 'Plot ID', 'Plot Name', 'Polli_type', 'Shelling Type', 'Generation', 'Field_Name', 'Source Stock', 'Row', 'Range', 'Plot', 'Block',
         'Rep', 'Kernel Num', 'Planting Date', 'Harvest Date', 'Comments'])
    for row in plot_loader:
        writer.writerow(
            [row.experiment.name, row.obs_plot.plot_id, row.obs_plot.plot_name, row.obs_plot.polli_type, row.obs_plot.get_shell_type(), row.obs_plot.gen, row.field.field_name, row.stock.seed_id,
             row.obs_plot.row_num, row.obs_plot.range_num, row.obs_plot.plot, row.obs_plot.block, row.obs_plot.rep,
             row.obs_plot.kernel_num, row.obs_plot.planting_date, row.obs_plot.harvest_date, row.obs_plot.comments])
    return response


@login_required
def download_field_map(request):
    plot_loader = sort_plot_loader(request)
    plot_objects = []  # Init: (range, row, experiment, plot_id)
    for obs in plot_loader:
        row_num = obs.obs_plot.row_num
        range_num = field_map_generator._get_column_letter(int(obs.obs_plot.range_num))

        plot_objects.append(field_map_generator.PlotCell(
            range_num=range_num, row_num=row_num, experiment=obs.experiment, plot_id=obs.obs_plot.plot_id,
            field=obs.experiment.field)
        )

    wb = field_map_generator.compile_info(plot_objects)
    response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="selected_experiment_maps.xlsx"'

    return response


@login_required
def download_field_map_by_field(request, field_id):
    plot_loader = ObsTracker.objects.filter(obs_entity_type='plot', field_id=field_id)
    field = Field.objects.get(id=field_id)

    if len(plot_loader) == 0:
        wb = field_map_generator.empty_field()
    else:
        plot_objects = []
        for obs in plot_loader:
            row_num = obs.obs_plot.row_num
            range_num = field_map_generator._get_column_letter(int(obs.obs_plot.range_num))

            plot_objects.append(field_map_generator.PlotCell(
                range_num=range_num, row_num=row_num, experiment=obs.experiment, plot_id=obs.obs_plot.plot_id,
                field=obs.field)
            )

        wb = field_map_generator.compile_info(plot_objects, field)
    response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="{}-map.xlsx"'.format(field.field_name)

    return response


@login_required
def download_field_map_experiment(request, experiment_name):
    plot_loader = find_plot_from_experiment(experiment_name)
    plot_objects = []
    for obs in plot_loader:
        row_num = obs.obs_plot.row_num
        range_num = field_map_generator._get_column_letter(int(obs.obs_plot.range_num))

        plot_objects.append(field_map_generator.PlotCell(
            range_num=range_num, row_num=row_num, experiment=obs.experiment, plot_id=obs.obs_plot.plot_id,
            field=obs.experiment.field)
        )

    wb = field_map_generator.compile_info(plot_objects)
    response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="selected_experiment_maps.xlsx"'

    return response


def suggest_plot_experiment(request):
    context = RequestContext(request)
    context_dict = {}
    plot_experiment_list = []
    starts_with = ''
    if request.method == 'GET':
        starts_with = request.GET['suggestion']
    else:
        starts_with = request.POST['suggestion']
    if starts_with:
        plot_experiment_list = ObsTracker.objects.filter(obs_entity_type='plot',
                                                         experiment__name__contains=starts_with).values(
            'experiment__name', 'experiment__field__field_name', 'experiment__field__id', 'experiment__id').distinct()[
                               :2000]
    else:
        plot_experiment_list = None
    context_dict = checkbox_session_variable_check(request)
    context_dict['plot_experiment_list'] = plot_experiment_list
    return render_to_response('lab/plot/plot_experiment_list.html', context_dict, context)


def select_plot_experiment(request):
    context = RequestContext(request)
    context_dict = {}
    plot_loader = []
    checkbox_plot_experiment_name_list = []
    checkbox_plot_experiment_list = request.POST.getlist('checkbox_plot_experiment')
    field_set = Field.objects.exclude(id=1)
    for plot_experiment in checkbox_plot_experiment_list:
        plots = ObsTracker.objects.filter(obs_entity_type='plot', experiment__id=plot_experiment)
        plot_loader = list(chain(plots, plot_loader))
    for experiment_id in checkbox_plot_experiment_list:
        experiment_name = Experiment.objects.filter(id=experiment_id).values('name')
        checkbox_plot_experiment_name_list = list(chain(experiment_name, checkbox_plot_experiment_name_list))
    request.session['checkbox_plot_experiment'] = checkbox_plot_experiment_name_list
    request.session['checkbox_plot_experiment_id_list'] = checkbox_plot_experiment_list
    context_dict = checkbox_session_variable_check(request)
    context_dict['field_set'] = field_set
    context_dict['form'] = DownloadFieldForm
    context_dict['plot_loader'] = plot_loader
    context_dict['logged_in_user'] = request.user.username
    return render_to_response('lab/plot/plot_data.html', context_dict, context)


def checkbox_plot_loader_clear(request):
    context = RequestContext(request)
    context_dict = {}
    del request.session['checkbox_row_experiment']
    del request.session['checkbox_plot_experiment_id_list']
    plot_loader = sort_plot_loader(request)
    context_dict = checkbox_session_variable_check(request)
    context_dict['plot_loader'] = plot_loader
    context_dict['logged_in_user'] = request.user.username
    return render_to_response('lab/plot/plot_data.html', context_dict, context)


def show_all_plot_experiment(request):
    context = RequestContext(request)
    context_dict = {}
    plot_experiment_list = ObsTracker.objects.filter(obs_entity_type='plot').exclude(experiment__name=None).values(
        'experiment__name', 'experiment__field__field_name', 'experiment__user__username', 'experiment__purpose',
        'experiment__field__id', 'experiment__id').distinct()[:2000]
    context_dict = checkbox_session_variable_check(request)
    context_dict['plot_experiment_list'] = plot_experiment_list
    return render_to_response('lab/plot/plot_experiment_list.html', context_dict, context)


def checkbox_session_variable_check(request):
    context_dict = {}
    if request.session.get('checkbox_pedigree', None):
        context_dict['checkbox_pedigree'] = request.session.get('checkbox_pedigree')
    if request.session.get('checkbox_taxonomy', None):
        context_dict['checkbox_taxonomy'] = request.session.get('checkbox_taxonomy')
    if request.session.get('checkbox_isolatestock_disease', None):
        context_dict['checkbox_isolatestock_disease'] = request.session.get('checkbox_isolatestock_disease')
    if request.session.get('checkbox_isolatestock_taxonomy', None):
        context_dict['checkbox_isolatestock_taxonomy'] = request.session.get('checkbox_isolatestock_taxonomy')
    if request.session.get('checkbox_row_experiment', None):
        context_dict['checkbox_row_experiment'] = request.session.get('checkbox_row_experiment')
    if request.session.get('checkbox_plant_experiment', None):
        context_dict['checkbox_plant_experiment'] = request.session.get('checkbox_plant_experiment')
    if request.session.get('checkbox_tissue_experiment', None):
        context_dict['checkbox_tissue_experiment'] = request.session.get('checkbox_tissue_experiment')
    if request.session.get('checkbox_plate_experiment', None):
        context_dict['checkbox_plate_experiment'] = request.session.get('checkbox_plate_experiment')
    if request.session.get('checkbox_well_experiment', None):
        context_dict['checkbox_well_experiment'] = request.session.get('checkbox_well_experiment')
    if request.session.get('checkbox_dna_experiment', None):
        context_dict['checkbox_dna_experiment'] = request.session.get('checkbox_dna_experiment')
    if request.session.get('checkbox_culture_experiment', None):
        context_dict['checkbox_culture_experiment'] = request.session.get('checkbox_culture_experiment')
    if request.session.get('checkbox_maize_experiment', None):
        context_dict['checkbox_maize_experiment'] = request.session.get('checkbox_maize_experiment')
    if request.session.get('checkbox_sample_experiment', None):
        context_dict['checkbox_sample_experiment'] = request.session.get('checkbox_sample_experiment')
    if request.session.get('checkbox_microbe_experiment', None):
        context_dict['checkbox_microbe_experiment'] = request.session.get('checkbox_microbe_experiment')
    if request.session.get('checkbox_env_experiment', None):
        context_dict['checkbox_env_experiment'] = request.session.get('checkbox_env_experiment')
    if request.session.get('checkbox_measurement_experiment', None):
        context_dict['checkbox_measurement_experiment'] = request.session.get('checkbox_measurement_experiment')
    if request.session.get('checkbox_measurement_parameter', None):
        context_dict['checkbox_measurement_parameter'] = request.session.get('checkbox_measurement_parameter')
    if request.session.get('checkbox_seedinv_parameters', None):
        context_dict['checkbox_seedinv_parameters'] = request.session.get('checkbox_seedinv_parameters')
    if request.session.get('checkbox_isolatestock_disease_names', None):
        context_dict['checkbox_isolatestock_disease_names'] = request.session.get('checkbox_isolatestock_disease_names')
    if request.session.get('checkbox_isolatestock_taxonomy_names', None):
        context_dict['checkbox_isolatestock_taxonomy_names'] = request.session.get(
            'checkbox_isolatestock_taxonomy_names')
    return context_dict


def get_obs_tracker(obs_type, obs_id):
    kwargs = {obs_type: obs_id}
    try:
        obs_tracker = ObsTracker.objects.filter(**kwargs)
    except ObsTracker.DoesNotExist:
        obs_tracker = None
    if obs_tracker is not None:
        for tracker in obs_tracker:
            tracker = make_obs_tracker_info(tracker)
    return obs_tracker


def make_obs_tracker_info(tracker):
    obs_entity_type = tracker.obs_entity_type
    if obs_entity_type == 'stock':
        try:
            obs_tracker_id_info = [tracker.stock.seed_id, obs_entity_type, tracker.stock_id]
        except Stock.DoesNotExist:
            obs_tracker_id_info = ['No Stock', obs_entity_type, 1]
    elif obs_entity_type == 'isolatestock':
        try:
            obs_tracker_id_info = [tracker.isolatestock.isolatestock_id, obs_entity_type, tracker.isolatestock_id]
        except IsolateStock.DoesNotExist:
            obs_tracker_id_info = ['No IsolateStock', obs_entity_type, 1]
    elif obs_entity_type == 'isolate':
        try:
            obs_tracker_id_info = [tracker.isolate.isolate_id, obs_entity_type, tracker.isolate_id]
        except Isolate.DoesNotExist:
            obs_tracker_id_info = ['No Isolate', obs_entity_type, 1]
    elif obs_entity_type == 'maize':
        try:
            obs_tracker_id_info = [tracker.maize_sample.maize_sample_id, obs_entity_type, tracker.maize_sample_id]
        except MaizeSample.DoesNotExist:
            obs_tracker_id_info = ['No Maize Sample', obs_entity_type, 1]
    elif obs_entity_type == 'plot':
        try:
            obs_tracker_id_info = [tracker.obs_plot.plot_id, obs_entity_type, tracker.obs_plot_id]
        except ObsPlot.DoesNotExist:
            obs_tracker_id_info = ['No Plot', obs_entity_type, 1]
    elif obs_entity_type == 'plant':
        try:
            obs_tracker_id_info = [tracker.obs_plant.plant_id, obs_entity_type, tracker.obs_plant_id]
        except ObsPlant.DoesNotExist:
            obs_tracker_id_info = ['No Plant', obs_entity_type, 1]
    elif obs_entity_type == 'culture':
        try:
            obs_tracker_id_info = [tracker.obs_culture.culture_id, obs_entity_type, tracker.obs_culture_id]
        except ObsCulture.DoesNotExist:
            obs_tracker_id_info = ['No Culture', obs_entity_type, 1]
    elif obs_entity_type == 'dna':
        try:
            obs_tracker_id_info = [tracker.obs_dna.dna_id, obs_entity_type, tracker.obs_dna_id]
        except ObsDNA.DoesNotExist:
            obs_tracker_id_info = ['No DNA', obs_entity_type, 1]
    elif obs_entity_type == 'microbe':
        try:
            obs_tracker_id_info = [tracker.obs_microbe.microbe_id, obs_entity_type, tracker.obs_microbe_id]
        except ObsMicrobe.DoesNotExist:
            obs_tracker_id_info = ['No Microbe', obs_entity_type, 1]
    elif obs_entity_type == 'plate':
        try:
            obs_tracker_id_info = [tracker.obs_plate.plate_id, obs_entity_type, tracker.obs_plate_id]
        except ObsPlate.DoesNotExist:
            obs_tracker_id_info = ['No Plate', obs_entity_type, 1]
    elif obs_entity_type == 'well':
        try:
            obs_tracker_id_info = [tracker.obs_well.well_id, obs_entity_type, tracker.obs_well_id]
        except ObsWell.DoesNotExist:
            obs_tracker_id_info = ['No Well', obs_entity_type, 1]
    elif obs_entity_type == 'tissue':
        try:
            obs_tracker_id_info = [tracker.obs_tissue.tissue_id, obs_entity_type, tracker.obs_tissue_id]
        except ObsTissue.DoesNotExist:
            obs_tracker_id_info = ['No Tissue', obs_entity_type, 1]
    elif obs_entity_type == 'culture':
        try:
            obs_tracker_id_info = [tracker.obs_culture.culture_id, obs_entity_type, tracker.obs_culture_id]
        except ObsCulture.DoesNotExist:
            obs_tracker_id_info = ['No Culture', obs_entity_type, 1]
    elif obs_entity_type == 'sample':
        try:
            obs_tracker_id_info = [tracker.obs_sample.sample_id, obs_entity_type, tracker.obs_sample_id]
        except ObsSample.DoesNotExist:
            obs_tracker_id_info = ['No Stock', obs_entity_type, 1]
    elif obs_entity_type == 'extract':
        try:
            obs_tracker_id_info = [tracker.obs_extract.extract_id, obs_entity_type, tracker.obs_extract_id]
        except ObsExtract.DoesNotExist:
            obs_tracker_id_info = ['No Extract', obs_entity_type, 1]
    elif obs_entity_type == 'maize':
        try:
            obs_tracker_id_info = [tracker.maize_sample.maize_id, obs_entity_type, tracker.maize_sample_id]
        except MaizeSample.DoesNotExist:
            obs_tracker_id_info = ['No Maize Sample', obs_entity_type, 1]
    elif obs_entity_type == 'experiment':
        try:
            obs_tracker_id_info = [tracker.experiment.name, obs_entity_type, tracker.experiment.name]
        except Experiment.DoesNotExist:
            obs_tracker_id_info = ['No Experiment', obs_entity_type, 1]
    else:
        obs_tracker_id_info = ['None', 'No Type', 1]

    tracker.obs_id = obs_tracker_id_info[0]
    tracker.obs_id_url = '/lab/%s/%s/' % (obs_tracker_id_info[1], obs_tracker_id_info[2])
    return tracker


def get_seed_collected_from_plot(obs_type, obs_id):
    obs_tracker_type = 'source_obs__%s' % (obs_type)
    kwargs = {obs_tracker_type: obs_id, 'target_obs__obs_entity_type': 'stock'}
    try:
        obs_source = ObsTrackerSource.objects.filter(**kwargs)
    except ObsTrackerSource.DoesNotExist:
        obs_source = None
    if obs_source is not None:
        for tracker in obs_source:
            tracker = make_obs_tracker_info(tracker.target_obs)
    return obs_source


def get_obs_measurements(obs_type, obs_id):
    obs_tracker_type = 'obs_tracker__%s' % (obs_type)
    kwargs = {obs_tracker_type: obs_id}
    try:
        obs_measurements = Measurement.objects.filter(**kwargs)
    except ObsTracker.DoesNotExist:
        obs_measurements = None
    if obs_measurements is not None:
        for measurement in obs_measurements:
            measurement = make_obs_tracker_info(measurement.obs_tracker)
    return obs_measurements


